#!/usr/bin/env python3
"""Generate a clean, idempotent Oracle import for CBI fee tariffs 1404.

Source: user-supplied cbi-fee-1404.xlsx (one worksheet, 239 physical rows).
The generator deliberately avoids inventing formulas when the workbook does not
provide structured calculation values. Such rows are imported as EXTERNAL_VALUE
and are listed in the review CSV.
"""
from __future__ import annotations

import argparse
import csv
import hashlib
import json
import re
import unicodedata
from collections import Counter, defaultdict
from dataclasses import dataclass, asdict
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any, Iterable

from openpyxl import load_workbook

SOURCE_REG_CODE = "CBI_FEE_1404_04_35500"
POLICY_CODE = "CBI_RIAL_ELECTRONIC_1404"
POLICY_VERSION_NO = "1404.1"
CLASSIFICATION_CODE = "CBI_1404"
CREATED_BY = "CBI_1404_IMPORT"
EFFECTIVE_FROM = "2025-05-04"
SOURCE_EXPECTED_SHA256 = "f43842c539427292dd86ce51562b2dbe38367b2cb10642abb164135fc31c9588"

# Clean source-derived section codes. These are internal technical codes only;
# Persian source headings remain unchanged in NAME_FA/DESCRIPTION.
SECTION_CODES = {
    "صدور ضمانتنامه شرکت در مناقصه , مزایده و سایر ضمانت نامه ها": ("GAR", "CBI_GUARANTEE_FEE_GROUP"),
    "حواله ها": ("REM", "CBI_REMITTANCE_FEE_GROUP"),
    "صندوق های اجاره ای": ("SDB", "CBI_SAFE_DEPOSIT_FEE_GROUP"),
    "مدیریت اوراق بهادار مشتریان (اوراق مشارکت , گواهی سپرده , صکوک و ...)": ("SEC", "CBI_SECURITIES_FEE_GROUP"),
    "وصول بروات": ("COL", "CBI_BILL_COLLECTION_FEE_GROUP"),
    "حساب قرض الحسنه جاری": ("CUR", "CBI_CURRENT_ACCOUNT_FEE_GROUP"),
    "حساب های قرض الحسنه پس انداز و سپرده کوتاه مدت": ("SAV", "CBI_SAVINGS_DEPOSIT_FEE_GROUP"),
    "صدور گواهی": ("CERT", "CBI_CERTIFICATE_FEE_GROUP"),
    "ارائه صورت حساب": ("STMT", "CBI_STATEMENT_FEE_GROUP"),
    "سایر خدمات انواع حساب": ("ACCSVC", "CBI_ACCOUNT_SERVICE_FEE_GROUP"),
    "ارزیابی اموال منقول و غیرمنقول": ("APP", "CBI_APPRAISAL_FEE_GROUP"),
    "اعتبارات": ("CRD", "CBI_CREDIT_FEE_GROUP"),
    "سایر خدمات": ("OTH", "CBI_OTHER_SERVICE_FEE_GROUP"),
    "خدمات الکترونیکی": ("ELEC", "CBI_ELECTRONIC_SERVICE_FEE_GROUP"),
    "خدمات الکترونیکی اعتبار اسنادی داخلی - ریالی": ("ELC", "CBI_ELECTRONIC_LC_RIAL_FEE_GROUP"),
    "خدمات الکترونیکی ضمانت نامه بانکی - ریالی": ("EGAR", "CBI_ELECTRONIC_GUARANTEE_RIAL_FEE_GROUP"),
    "خدمات الکترونیکی برات الکترونیک": ("EBILL", "CBI_ELECTRONIC_BILL_FEE_GROUP"),
}

# Existing rich baseline definitions that already model source rules in more detail.
# Reuse these FEE_CODE values instead of creating duplicate official definitions.
EXISTING_FEE_CODE_BY_SOURCE_ROW = {
    3: "CBI_GUARANTEE_DEPOSIT_FEE",       # tariff 1-2
    42: "CBI_BILL_COLLECTION_FEE",         # tariff 5-1
    54: "CBI_CHEQUEBOOK_ISSUE_FEE",        # tariff 6-1-9
    90: "CBI_PROPERTY_APPRAISAL_FEE",       # tariff 7-2 tier group
    118: "CBI_UNUSED_LIMIT_COMMITMENT_FEE", # tariff 8-15
    119: "CBI_CREDIT_EXPERTISE_FEE",        # tariff 8-16
}

# Five physical 3-row bracket groups are one logical tariff each.
TIER_GROUPS = {
    (87, 88, 89): {
        "name": "ارزیابی ماشین آلات و کالا",
        "first_upper": Decimal("2000000000"),
        "tariff_code": "7-1",
    },
    (90, 91, 92): {
        "name": "ارزیابی املاک و ساختمان",
        "first_upper": Decimal("20000000000"),
        "tariff_code": "7-2",
    },
    (93, 94, 95): {
        "name": "ارزیابی املاک مزروعی و باغات",
        "first_upper": Decimal("2000000000"),
        "tariff_code": "7-3",
    },
    (96, 97, 98): {
        "name": "ارزیابی تسهیلات مشارکت مدنی و پیشرفت فیزیکی پروژه ها (برآورد هزینه و آورده غیرنقدی شریک)",
        "first_upper": Decimal("2000000000"),
        "tariff_code": "7-3",
    },
    (101, 102, 103): {
        "name": "ارزیابی وثایق غیرمنقول مازاد",
        "first_upper": Decimal("2000000000"),
        "tariff_code": "7-3",
    },
}
TIER_ROW_SET = {r for group in TIER_GROUPS for r in group}

# Known source conflict: structured amount and prose amount disagree by a factor of 10.
# Do not choose one silently; import the rule as EXTERNAL_VALUE and expose the conflict.
CONFLICT_ROWS = {
    45: "مبلغ ساختاری 75,000 ریال است ولی متن نوع خدمت 7,500 ریال برای هر برگ ذکر می‌کند؛ نیازمند تأیید منبع مقرراتی.",
    60: "مبلغ ساختاری 75,000 ریال است ولی متن نوع خدمت 7,500 ریال بابت هر برگ چک ذکر می‌کند؛ نیازمند تأیید منبع مقرراتی.",
    154: "مبلغ ساختاری 18,000 ریال است ولی متن نوع خدمت 180,000 ریال ذکر می‌کند؛ نیازمند تأیید منبع مقرراتی.",
}

@dataclass
class SourceRow:
    excel_row: int
    section: str
    tariff_code: str | None
    service_code: str | None
    name_fa: str
    calc_type: str | None
    amount: Decimal | None
    percent: Decimal | None
    min_fee: Decimal | None
    max_fee: Decimal | None
    period: str | None
    notes: str | None
    service_text: str | None
    max_text: str | None

@dataclass
class LogicalTariff:
    logical_no: int
    source_rows: list[int]
    section: str
    section_code: str
    feature_code: str
    tariff_code: str | None
    service_code: str | None
    fee_code: str
    name_fa: str
    calc_source_type: str | None
    calc_strategy: str
    basis_type: str
    fixed_amount: Decimal | None
    source_percent: Decimal | None
    rate_value: Decimal | None
    min_fee: Decimal | None
    max_fee: Decimal | None
    rate_period_code: str | None
    currency_code: str
    source_notes: str | None
    source_service_text: str | None
    source_max_text: str | None
    import_status: str
    review_reason: str | None
    config_hash: str
    tier_first_upper: Decimal | None = None


def clean_text(value: Any) -> str | None:
    if value is None:
        return None
    s = str(value)
    s = unicodedata.normalize("NFKC", s)
    trans = str.maketrans({
        "ي": "ی", "ى": "ی", "ك": "ک", "ة": "ه", "ۀ": "ه",
        "‌": " ", "‍": " ", "\u200e": " ", "\u200f": " ",
        "\ufeff": " ", "\xa0": " ", "ـ": "",
    })
    s = s.translate(trans)
    s = s.replace("\r", " ").replace("\n", " ").replace("\t", " ")
    s = re.sub(r"\s+", " ", s).strip()
    # Normalize punctuation spacing without altering the actual wording.
    s = re.sub(r"\s+,\s*", " , ", s)
    s = re.sub(r"\s*؛\s*", "؛ ", s)
    s = re.sub(r"\(\s+", "(", s)
    s = re.sub(r"\s+\)", ")", s)
    return s or None


def decimal_value(value: Any) -> Decimal | None:
    if value is None or value == "":
        return None
    try:
        return Decimal(str(value))
    except InvalidOperation as e:
        raise ValueError(f"Non-numeric source value: {value!r}") from e


def tariff_code_value(value: Any) -> str | None:
    if value is None or value == "":
        return None
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return clean_text(value)


def canonical_section(value: str) -> str:
    # Apply only orthographic cleanup needed to match internal section mapping.
    s = clean_text(value) or ""
    s = s.replace("ضمانتنامه", "ضمانتنامه")
    s = s.replace("مزايده", "مزایده").replace("ساير", "سایر")
    s = s.replace("مديريت", "مدیریت").replace("مشتريان", "مشتریان")
    s = s.replace("ارزيابی", "ارزیابی").replace("ارزيابي", "ارزیابی")
    s = s.replace("غيرمنقول", "غیرمنقول")
    s = s.replace("الکترونيکی", "الکترونیکی").replace("الکترونيک", "الکترونیک")
    s = s.replace("ريالی", "ریالی").replace("ريالي", "ریالی")
    return s


def sql_str(value: str | None) -> str:
    if value is None:
        return "NULL"
    return "'" + value.replace("'", "''") + "'"


def sql_num(value: Decimal | None) -> str:
    if value is None:
        return "NULL"
    # Avoid scientific notation in Oracle seed SQL.
    s = format(value, "f")
    if "." in s:
        s = s.rstrip("0").rstrip(".")
    return s or "0"


def build_description(item: LogicalTariff) -> str:
    parts = [
        f"منبع: بخشنامه کارمزد بانک مرکزی ۱۴۰۴؛ سرفصل: {item.section}",
        f"ردیف(های) فایل Excel: {','.join(map(str, item.source_rows))}",
    ]
    if item.tariff_code:
        parts.append(f"ردیف تعرفه: {item.tariff_code}")
    if item.service_code:
        parts.append(f"کد خدمت/پیام: {item.service_code}")
    if item.source_service_text:
        parts.append(f"نوع خدمت/فرمول منبع: {item.source_service_text}")
    if item.source_notes:
        parts.append(f"توضیحات منبع: {item.source_notes}")
    if item.source_max_text:
        parts.append(f"شرح حداکثر/عنوان تکمیلی منبع: {item.source_max_text}")
    if item.review_reason:
        parts.append(f"نیازمند بازبینی: {item.review_reason}")
    text = " | ".join(parts)
    return text[:1990]


def normalize_name(name: str) -> str:
    s = clean_text(name) or ""
    # Safe orthographic normalization only; no semantic rewriting.
    for a, b in {
        "ارزيابي": "ارزیابی", "ارزيابی": "ارزیابی", "املاک": "املاک",
        "تسهيلات": "تسهیلات", "حقيقی": "حقیقی", "حقوقی": "حقوقی",
        "مشتري": "مشتری", "مشتری": "مشتری", "سپرده مدت دار": "سپرده مدت دار",
        "ضمانت نامه": "ضمانت نامه", "ريال": "ریال", "ريالی": "ریالی",
        "غير": "غیر", "حسابهاي": "حساب های", "کليه": "کلیه",
    }.items():
        s = s.replace(a, b)
    return s.strip()


def source_rows_from_workbook(path: Path) -> tuple[list[SourceRow], str]:
    raw = path.read_bytes()
    sha = hashlib.sha256(raw).hexdigest()
    wb = load_workbook(path, data_only=True, read_only=False)
    if wb.sheetnames != ["کل کارمزدها"]:
        raise ValueError(f"Unexpected workbook sheets: {wb.sheetnames}")
    ws = wb["کل کارمزدها"]
    expected_headers = [
        "سرفصل", "ردیف", "کد", "شرح کارمزد", "نوع محاسبه کارمزد", "مبلغ", "درصد",
        "حداقل کارمزد", "حداکثر کارمزد", "دوره", "توضيحات", "نوع خدمت", None, "حداکثر میزان کارمزد"
    ]
    actual = [ws.cell(1, c).value for c in range(1, 15)]
    if actual != expected_headers:
        raise ValueError(f"Unexpected workbook headers: {actual}")
    if ws.max_row != 240:
        raise ValueError(f"Expected 239 source rows + header, got max_row={ws.max_row}")

    rows: list[SourceRow] = []
    last_section: str | None = None
    for r in range(2, ws.max_row + 1):
        section_raw = clean_text(ws.cell(r, 1).value)
        if section_raw:
            last_section = canonical_section(section_raw)
        if not last_section:
            raise ValueError(f"Cannot resolve section at row {r}")
        section = last_section
        if section not in SECTION_CODES:
            raise ValueError(f"Unknown normalized section at row {r}: {section!r}")
        name = normalize_name(ws.cell(r, 4).value)
        if not name:
            raise ValueError(f"Missing fee description at row {r}")
        rows.append(SourceRow(
            excel_row=r,
            section=section,
            tariff_code=tariff_code_value(ws.cell(r, 2).value),
            service_code=clean_text(ws.cell(r, 3).value),
            name_fa=name,
            calc_type=clean_text(ws.cell(r, 5).value),
            amount=decimal_value(ws.cell(r, 6).value),
            percent=decimal_value(ws.cell(r, 7).value),
            min_fee=decimal_value(ws.cell(r, 8).value),
            max_fee=decimal_value(ws.cell(r, 9).value),
            period=clean_text(ws.cell(r, 10).value),
            notes=clean_text(ws.cell(r, 11).value),
            service_text=clean_text(ws.cell(r, 12).value),
            max_text=clean_text(ws.cell(r, 14).value),
        ))
    return rows, sha


def rate_strategy(row: SourceRow) -> tuple[str, str, Decimal | None, Decimal | None, Decimal | None, Decimal | None, str | None, str, str | None]:
    """Return strategy/basis/fixed/source_pct/rate/min/max/period/status/review."""
    if row.excel_row in CONFLICT_ROWS:
        return "EXTERNAL_VALUE", "Other", None, row.percent, None, None, None, None, "REVIEW_CONFLICT", CONFLICT_ROWS[row.excel_row]

    typ = row.calc_type
    if typ == "مبلغ ثابت":
        if row.amount is None:
            return "EXTERNAL_VALUE", "Other", None, row.percent, None, None, None, None, "REVIEW_SOURCE_GAP", "نوع محاسبه مبلغ ثابت است ولی ستون مبلغ خالی است."
        if row.percent is not None or row.min_fee is not None or row.max_fee is not None:
            # Any mixed numeric semantics must be reviewed rather than silently interpreted.
            return "EXTERNAL_VALUE", "Other", None, row.percent, None, None, None, None, "REVIEW_CONFLICT", "برای مبلغ ثابت، ستون های نرخ/حداقل/حداکثر نیز مقدار دارند؛ محاسبه اجرایی تا تأیید منبع غیرفعال شد."
        return "FIXED", "Flat", row.amount, None, None, None, None, None, "SAFE", None

    if typ in ("درصد", "درصد مبلغ با حداقل/حداکثر"):
        if row.percent is None:
            return "EXTERNAL_VALUE", "Other", None, None, None, None, None, None, "REVIEW_SOURCE_GAP", "نوع محاسبه درصدی است ولی ستون درصد خالی است."
        rate = row.percent / Decimal("100")
        period = "YEAR" if row.period == "سال" else None
        if period:
            strategy = "ANNUALIZED_PERCENTAGE"
            basis = "RateWithMinimumAmount" if row.min_fee is not None else "Percentage"
        elif row.min_fee is not None and row.max_fee is not None:
            strategy = "PERCENTAGE_FLOOR_CAP"
            basis = "Percentage"
        elif row.min_fee is not None:
            strategy = "PERCENTAGE_WITH_FLOOR"
            basis = "RateWithMinimumAmount"
        elif row.max_fee is not None:
            strategy = "PERCENTAGE_WITH_CAP"
            basis = "RateWithMaximumAmount"
        else:
            strategy = "PERCENTAGE"
            basis = "Percentage"
        return strategy, basis, None, row.percent, rate, row.min_fee, row.max_fee, period, "SAFE", None

    if typ == "پلکانی":
        # Standalone 'tiered' row without the structured 3-row bracket pattern is external.
        return "EXTERNAL_VALUE", "Other", None, row.percent, None, None, None, None, "EXTERNAL_RULE", "فایل منبع نوع را پلکانی اعلام کرده ولی پله های ساختاری قابل اتکا در همان ردیف وجود ندارد."

    # Blank calculation type: only a source-explicit free service is safe to encode as zero.
    service = (row.service_text or "") + " " + (row.notes or "")
    if "بدون کارمزد" in service and row.amount is None and row.percent is None:
        return "FIXED", "Flat", Decimal("0"), None, None, None, None, None, "SAFE", None

    return "EXTERNAL_VALUE", "Other", None, row.percent, None, None, None, None, "EXTERNAL_RULE", "فرمول ساختاری در ستون های فایل منبع ارائه نشده است؛ متن منبع بدون تبدیل به فرمول اجرایی حفظ شد."


def make_fee_code(section_short: str, tariff_code: str | None, occurrence: int, first_row: int) -> str:
    if tariff_code:
        token = re.sub(r"[^A-Za-z0-9]+", "_", tariff_code.upper()).strip("_") or f"SRC{first_row}"
    else:
        token = f"SRC{first_row}"
    code = f"CBI1404_{section_short}_{token}_{occurrence:02d}"
    return code[:80]


def build_logical(rows: list[SourceRow]) -> list[LogicalTariff]:
    by_excel = {r.excel_row: r for r in rows}
    logical_raw: list[dict[str, Any]] = []

    # Add tier groups first at the physical position of their first source row.
    group_by_first = {group[0]: (group, cfg) for group, cfg in TIER_GROUPS.items()}
    skip = set()
    for row in rows:
        if row.excel_row in skip:
            continue
        if row.excel_row in group_by_first:
            group, cfg = group_by_first[row.excel_row]
            group_rows = [by_excel[n] for n in group]
            skip.update(group)
            first = group_rows[0]
            service_text = next((r.service_text for r in group_rows if r.service_text), None)
            notes = " | ".join(filter(None, (r.notes for r in group_rows))) or None
            max_text = " | ".join(filter(None, (r.max_text for r in group_rows))) or None
            logical_raw.append({
                "source_rows": list(group),
                "section": first.section,
                "tariff_code": cfg["tariff_code"],
                "service_code": first.service_code,
                "name_fa": normalize_name(cfg["name"]),
                "calc_source_type": "پلکانی",
                "strategy": "COMPOSITE",
                "basis": "RatePerBrackets",
                "fixed": None,
                "source_percent": Decimal("0.2"),
                "rate": Decimal("0.002"),
                "min_fee": None,
                "max_fee": None,
                "period": None,
                "source_notes": notes,
                "service_text": service_text,
                "max_text": max_text,
                "status": "SAFE_TIERED_EXTERNAL",
                "review": "پله سوم و تعرفه کارشناس رسمی مقدار بیرونی هستند؛ وابستگی خارجی صریحاً مدل شده و عددی جعل نشده است.",
                "tier_first_upper": cfg["first_upper"],
            })
            continue
        if row.excel_row in TIER_ROW_SET:
            continue
        strategy, basis, fixed, source_pct, rate, min_fee, max_fee, period, status, review = rate_strategy(row)
        logical_raw.append({
            "source_rows": [row.excel_row],
            "section": row.section,
            "tariff_code": row.tariff_code,
            "service_code": row.service_code,
            "name_fa": row.name_fa,
            "calc_source_type": row.calc_type,
            "strategy": strategy,
            "basis": basis,
            "fixed": fixed,
            "source_percent": source_pct,
            "rate": rate,
            "min_fee": min_fee,
            "max_fee": max_fee,
            "period": period,
            "source_notes": row.notes,
            "service_text": row.service_text,
            "max_text": row.max_text,
            "status": status,
            "review": review,
            "tier_first_upper": None,
        })

    logical_raw.sort(key=lambda x: x["source_rows"][0])
    # Occurrence is per section+external tariff code. It preserves duplicate CBI row codes without collisions.
    occurrence: defaultdict[tuple[str, str], int] = defaultdict(int)
    result: list[LogicalTariff] = []
    for idx, raw in enumerate(logical_raw, start=1):
        section_short, feature_code = SECTION_CODES[raw["section"]]
        tariff_key = raw["tariff_code"] or f"SRC{raw['source_rows'][0]}"
        occurrence[(section_short, tariff_key)] += 1
        fee_code = EXISTING_FEE_CODE_BY_SOURCE_ROW.get(raw["source_rows"][0])
        if not fee_code:
            fee_code = make_fee_code(section_short, raw["tariff_code"], occurrence[(section_short, tariff_key)], raw["source_rows"][0])
        payload = "|".join([
            fee_code, raw["section"], raw["tariff_code"] or "", raw["name_fa"], raw["strategy"],
            str(raw["fixed"] or ""), str(raw["rate"] or ""), str(raw["min_fee"] or ""), str(raw["max_fee"] or ""),
            ",".join(map(str, raw["source_rows"])),
        ])
        config_hash = hashlib.sha256(payload.encode("utf-8")).hexdigest()
        status = raw["status"]
        if raw["source_rows"][0] in EXISTING_FEE_CODE_BY_SOURCE_ROW:
            status = "EXISTING_ENRICHED"
        result.append(LogicalTariff(
            logical_no=idx,
            source_rows=raw["source_rows"],
            section=raw["section"],
            section_code=section_short,
            feature_code=feature_code,
            tariff_code=raw["tariff_code"],
            service_code=raw["service_code"],
            fee_code=fee_code,
            name_fa=raw["name_fa"],
            calc_source_type=raw["calc_source_type"],
            calc_strategy=raw["strategy"],
            basis_type=raw["basis"],
            fixed_amount=raw["fixed"],
            source_percent=raw["source_percent"],
            rate_value=raw["rate"],
            min_fee=raw["min_fee"],
            max_fee=raw["max_fee"],
            rate_period_code=raw["period"],
            currency_code="IRR",
            source_notes=raw["source_notes"],
            source_service_text=raw["service_text"],
            source_max_text=raw["max_text"],
            import_status=status,
            review_reason=raw["review"],
            config_hash=config_hash,
            tier_first_upper=raw["tier_first_upper"],
        ))
    return result


def csv_value(value: Any) -> Any:
    if isinstance(value, Decimal):
        return format(value, "f")
    if isinstance(value, list):
        return ",".join(map(str, value))
    return value


def write_csv(path: Path, items: list[LogicalTariff]) -> None:
    fields = [
        "logical_no", "source_rows", "section", "section_code", "feature_code", "tariff_code", "service_code",
        "fee_code", "name_fa", "calc_source_type", "calc_strategy", "basis_type", "fixed_amount", "source_percent",
        "rate_value", "min_fee", "max_fee", "rate_period_code", "currency_code", "source_notes", "source_service_text",
        "source_max_text", "import_status", "review_reason", "config_hash", "tier_first_upper"
    ]
    with path.open("w", encoding="utf-8-sig", newline="") as f:
        w = csv.DictWriter(f, fieldnames=fields)
        w.writeheader()
        for item in items:
            d = asdict(item)
            w.writerow({k: csv_value(d[k]) for k in fields})


def write_review_csv(path: Path, items: list[LogicalTariff]) -> None:
    review = [x for x in items if x.import_status not in {"SAFE", "EXISTING_ENRICHED"}]
    write_csv(path, review)


def scalar_def_version(fee_code: str) -> str:
    return f"(SELECT v.FEE_DEFINITION_VERSION_ID FROM FEE_DEFINITION_VERSION v JOIN FEE_DEFINITION d ON d.FEE_DEFINITION_ID=v.FEE_DEFINITION_ID WHERE d.FEE_CODE={sql_str(fee_code)} AND v.VERSION_NO='1.0')"


def scalar_calc_rule(fee_code: str, rule_code: str) -> str:
    return f"(SELECT r.CALCULATION_RULE_ID FROM FEE_CALCULATION_RULE r WHERE r.FEE_DEFINITION_VERSION_ID={scalar_def_version(fee_code)} AND r.RULE_CODE={sql_str(rule_code)})"


def merge_feature(section: str, short: str, feature_code: str) -> str:
    return f"""MERGE INTO FEE_FEATURE t
USING (SELECT {sql_str(feature_code)} FEATURE_CODE, {sql_str(section)} NAME_FA, 'Other' FEE_TYPE_CODE, {sql_str('CBI_' + short)} DEFAULT_CATEGORY_CODE FROM dual) s
ON (t.FEATURE_CODE=s.FEATURE_CODE)
WHEN MATCHED THEN UPDATE SET t.NAME_FA=s.NAME_FA, t.FEE_TYPE_CODE=s.FEE_TYPE_CODE, t.DEFAULT_CATEGORY_CODE=s.DEFAULT_CATEGORY_CODE,
    t.DESCRIPTION='گروه تعرفه برگرفته از سرفصل فایل بانک مرکزی ۱۴۰۴', t.IS_ACTIVE='Y', t.UPDATED_AT=SYSTIMESTAMP, t.UPDATED_BY='{CREATED_BY}', t.RECORD_VERSION=t.RECORD_VERSION+1
WHEN NOT MATCHED THEN INSERT (FEATURE_CODE,NAME_FA,FEE_TYPE_CODE,DEFAULT_CATEGORY_CODE,DESCRIPTION,IS_ACTIVE,CREATED_BY)
VALUES (s.FEATURE_CODE,s.NAME_FA,s.FEE_TYPE_CODE,s.DEFAULT_CATEGORY_CODE,'گروه تعرفه برگرفته از سرفصل فایل بانک مرکزی ۱۴۰۴','Y','{CREATED_BY}');
"""


def merge_definition(item: LogicalTariff) -> str:
    desc = build_description(item)
    return f"""MERGE INTO FEE_DEFINITION t
USING (SELECT (SELECT FEE_FEATURE_ID FROM FEE_FEATURE WHERE FEATURE_CODE={sql_str(item.feature_code)}) FEE_FEATURE_ID,
              {sql_str(item.fee_code)} FEE_CODE, {sql_str(item.name_fa)} NAME_FA, {sql_str('CBI_' + item.section_code)} CATEGORY_CODE,
              {sql_str(CLASSIFICATION_CODE)} CLASSIFICATION_CODE, {sql_str(desc)} DESCRIPTION FROM dual) s
ON (t.FEE_CODE=s.FEE_CODE)
WHEN MATCHED THEN UPDATE SET t.FEE_FEATURE_ID=s.FEE_FEATURE_ID, t.NAME_FA=s.NAME_FA, t.CATEGORY_CODE=s.CATEGORY_CODE,
    t.CLASSIFICATION_CODE=s.CLASSIFICATION_CODE, t.DESCRIPTION=s.DESCRIPTION, t.IS_ACTIVE='Y',
    t.UPDATED_AT=SYSTIMESTAMP, t.UPDATED_BY='{CREATED_BY}', t.RECORD_VERSION=t.RECORD_VERSION+1
WHEN NOT MATCHED THEN INSERT (FEE_FEATURE_ID,FEE_CODE,NAME_FA,CATEGORY_CODE,CLASSIFICATION_CODE,DESCRIPTION,IS_ACTIVE,CREATED_BY)
VALUES (s.FEE_FEATURE_ID,s.FEE_CODE,s.NAME_FA,s.CATEGORY_CODE,s.CLASSIFICATION_CODE,s.DESCRIPTION,'Y','{CREATED_BY}');
"""


def merge_version(item: LogicalTariff) -> str:
    reason = build_description(item)
    return f"""MERGE INTO FEE_DEFINITION_VERSION t
USING (SELECT d.FEE_DEFINITION_ID,
              (SELECT pv.POLICY_VERSION_ID FROM FEE_POLICY_VERSION pv JOIN FEE_POLICY_SET ps ON ps.POLICY_SET_ID=pv.POLICY_SET_ID WHERE ps.POLICY_CODE={sql_str(POLICY_CODE)} AND pv.VERSION_NO={sql_str(POLICY_VERSION_NO)}) POLICY_VERSION_ID,
              (SELECT REGULATORY_SOURCE_ID FROM FEE_REGULATORY_SOURCE WHERE SOURCE_CODE={sql_str(SOURCE_REG_CODE)}) REGULATORY_SOURCE_ID,
              {sql_str(item.tariff_code)} REGULATORY_TARIFF_CODE, {sql_str(item.name_fa)} FEE_PLAN_NAME,
              {sql_str(reason)} FEE_REASON, {sql_str(item.config_hash)} CONFIG_HASH
       FROM FEE_DEFINITION d WHERE d.FEE_CODE={sql_str(item.fee_code)}) s
ON (t.FEE_DEFINITION_ID=s.FEE_DEFINITION_ID AND t.VERSION_NO='1.0')
WHEN MATCHED THEN UPDATE SET t.POLICY_VERSION_ID=s.POLICY_VERSION_ID, t.REGULATORY_SOURCE_ID=s.REGULATORY_SOURCE_ID,
    t.REGULATORY_TARIFF_CODE=s.REGULATORY_TARIFF_CODE, t.STATUS_CODE='ACTIVE', t.FEE_PLAN_NAME=s.FEE_PLAN_NAME,
    t.FEE_REASON=s.FEE_REASON, t.EFFECTIVE_FROM=DATE '{EFFECTIVE_FROM}', t.CONFIG_HASH=s.CONFIG_HASH,
    t.UPDATED_AT=SYSTIMESTAMP, t.UPDATED_BY='{CREATED_BY}', t.RECORD_VERSION=t.RECORD_VERSION+1
WHEN NOT MATCHED THEN INSERT (FEE_DEFINITION_ID,POLICY_VERSION_ID,REGULATORY_SOURCE_ID,REGULATORY_TARIFF_CODE,VERSION_NO,VERSION_TYPE_CODE,STATUS_CODE,FEE_PLAN_NAME,FEE_REASON,EFFECTIVE_FROM,CONFIG_HASH,CREATED_BY)
VALUES (s.FEE_DEFINITION_ID,s.POLICY_VERSION_ID,s.REGULATORY_SOURCE_ID,s.REGULATORY_TARIFF_CODE,'1.0','MAJOR','ACTIVE',s.FEE_PLAN_NAME,s.FEE_REASON,DATE '{EFFECTIVE_FROM}',s.CONFIG_HASH,'{CREATED_BY}');
"""


def merge_calc_rule(item: LogicalTariff, rule_code: str = "CBI1404_SOURCE", strategy: str | None = None,
                    basis: str | None = None, fixed: Decimal | None = None, rate: Decimal | None = None,
                    min_fee: Decimal | None = None, max_fee: Decimal | None = None, period: str | None = None,
                    name: str | None = None, description: str | None = None, priority: int = 10) -> str:
    strategy = strategy or item.calc_strategy
    basis = basis or item.basis_type
    if fixed is None and strategy == item.calc_strategy:
        fixed = item.fixed_amount
    if rate is None and strategy == item.calc_strategy:
        rate = item.rate_value
    if min_fee is None and strategy == item.calc_strategy:
        min_fee = item.min_fee
    if max_fee is None and strategy == item.calc_strategy:
        max_fee = item.max_fee
    if period is None and strategy == item.calc_strategy:
        period = item.rate_period_code
    name = name or f"قاعده تعرفه بانک مرکزی - {item.name_fa}"
    description = (description or build_description(item))[:1990]
    return f"""MERGE INTO FEE_CALCULATION_RULE t
USING (SELECT {scalar_def_version(item.fee_code)} FEE_DEFINITION_VERSION_ID FROM dual) s
ON (t.FEE_DEFINITION_VERSION_ID=s.FEE_DEFINITION_VERSION_ID AND t.RULE_CODE={sql_str(rule_code)})
WHEN MATCHED THEN UPDATE SET t.NAME_FA={sql_str(name)}, t.PRIORITY_NO={priority}, t.CALCULATION_STRATEGY_CODE={sql_str(strategy)},
    t.BASIS_TYPE_CODE={sql_str(basis)}, t.FIXED_AMOUNT={sql_num(fixed)}, t.RATE_VALUE={sql_num(rate)},
    t.MIN_FEE_AMOUNT={sql_num(min_fee)}, t.MAX_FEE_AMOUNT={sql_num(max_fee)}, t.RATE_PERIOD_CODE={sql_str(period)},
    t.CURRENCY_CODE='IRR', t.ROUNDING_MODE_CODE='HALF_UP', t.ROUNDING_SCALE=0, t.EFFECTIVE_FROM=DATE '{EFFECTIVE_FROM}',
    t.IS_ACTIVE='Y', t.DESCRIPTION={sql_str(description)}, t.UPDATED_AT=SYSTIMESTAMP, t.UPDATED_BY='{CREATED_BY}', t.RECORD_VERSION=t.RECORD_VERSION+1
WHEN NOT MATCHED THEN INSERT (FEE_DEFINITION_VERSION_ID,RULE_CODE,NAME_FA,PRIORITY_NO,CALCULATION_STRATEGY_CODE,BASIS_TYPE_CODE,FIXED_AMOUNT,RATE_VALUE,MIN_FEE_AMOUNT,MAX_FEE_AMOUNT,RATE_PERIOD_CODE,CURRENCY_CODE,ROUNDING_MODE_CODE,ROUNDING_SCALE,EFFECTIVE_FROM,IS_ACTIVE,DESCRIPTION,CREATED_BY)
VALUES (s.FEE_DEFINITION_VERSION_ID,{sql_str(rule_code)},{sql_str(name)},{priority},{sql_str(strategy)},{sql_str(basis)},{sql_num(fixed)},{sql_num(rate)},{sql_num(min_fee)},{sql_num(max_fee)},{sql_str(period)},'IRR','HALF_UP',0,DATE '{EFFECTIVE_FROM}','Y',{sql_str(description)},'{CREATED_BY}');
"""


def merge_input(fee_code: str, rule_code: str, input_code: str, name_fa: str, order: int) -> str:
    calc = scalar_calc_rule(fee_code, rule_code)
    return f"""MERGE INTO FEE_INPUT_DEFINITION t
USING (SELECT {calc} CALCULATION_RULE_ID FROM dual) s
ON (t.CALCULATION_RULE_ID=s.CALCULATION_RULE_ID AND t.INPUT_CODE={sql_str(input_code)})
WHEN MATCHED THEN UPDATE SET t.NAME_FA={sql_str(name_fa)}, t.DATA_TYPE_CODE='NUMBER', t.UNIT_CODE='IRR', t.MANDATORY_FLAG='Y', t.DISPLAY_ORDER={order},
    t.UPDATED_AT=SYSTIMESTAMP, t.UPDATED_BY='{CREATED_BY}', t.RECORD_VERSION=t.RECORD_VERSION+1
WHEN NOT MATCHED THEN INSERT (CALCULATION_RULE_ID,INPUT_CODE,NAME_FA,DATA_TYPE_CODE,UNIT_CODE,MANDATORY_FLAG,DISPLAY_ORDER,CREATED_BY)
VALUES (s.CALCULATION_RULE_ID,{sql_str(input_code)},{sql_str(name_fa)},'NUMBER','IRR','Y',{order},'{CREATED_BY}');
"""


def merge_component(fee_code: str, rule_code: str, seq: int, node_type: str, description: str,
                    parent_seq: int | None = None, operator: str | None = None, input_code: str | None = None,
                    constant: Decimal | None = None, reference: str | None = None) -> str:
    calc = scalar_calc_rule(fee_code, rule_code)
    parent = "NULL" if parent_seq is None else f"(SELECT p.RULE_COMPONENT_ID FROM FEE_RULE_COMPONENT p WHERE p.CALCULATION_RULE_ID={calc} AND p.SEQUENCE_NO={parent_seq})"
    return f"""MERGE INTO FEE_RULE_COMPONENT t
USING (SELECT {calc} CALCULATION_RULE_ID FROM dual) s
ON (t.CALCULATION_RULE_ID=s.CALCULATION_RULE_ID AND t.SEQUENCE_NO={seq})
WHEN MATCHED THEN UPDATE SET t.PARENT_RULE_COMPONENT_ID={parent}, t.NODE_TYPE_CODE={sql_str(node_type)}, t.OPERATOR_CODE={sql_str(operator)},
    t.INPUT_CODE={sql_str(input_code)}, t.CONSTANT_NUMBER={sql_num(constant)}, t.REFERENCE_CODE={sql_str(reference)}, t.DESCRIPTION={sql_str(description)},
    t.UPDATED_AT=SYSTIMESTAMP, t.UPDATED_BY='{CREATED_BY}', t.RECORD_VERSION=t.RECORD_VERSION+1
WHEN NOT MATCHED THEN INSERT (CALCULATION_RULE_ID,PARENT_RULE_COMPONENT_ID,SEQUENCE_NO,NODE_TYPE_CODE,OPERATOR_CODE,INPUT_CODE,CONSTANT_NUMBER,REFERENCE_CODE,DESCRIPTION,CREATED_BY)
VALUES (s.CALCULATION_RULE_ID,{parent},{seq},{sql_str(node_type)},{sql_str(operator)},{sql_str(input_code)},{sql_num(constant)},{sql_str(reference)},{sql_str(description)},'{CREATED_BY}');
"""


def merge_tier(fee_code: str, rule_code: str, tier_no: int, name_fa: str, lower: Decimal, upper: Decimal | None,
               tier_basis: str, tier_strategy: str, fixed: Decimal | None = None, rate: Decimal | None = None) -> str:
    calc = scalar_calc_rule(fee_code, rule_code)
    return f"""MERGE INTO FEE_CALCULATION_TIER t
USING (SELECT {calc} CALCULATION_RULE_ID FROM dual) s
ON (t.CALCULATION_RULE_ID=s.CALCULATION_RULE_ID AND t.TIER_NO={tier_no})
WHEN MATCHED THEN UPDATE SET t.TIER_NAME_FA={sql_str(name_fa)}, t.LOWER_BOUND={sql_num(lower)}, t.UPPER_BOUND={sql_num(upper)},
    t.BOUND_UNIT_CODE='IRR', t.TIER_BASIS_CODE={sql_str(tier_basis)}, t.TIER_STRATEGY_CODE={sql_str(tier_strategy)},
    t.FIXED_AMOUNT={sql_num(fixed)}, t.RATE_VALUE={sql_num(rate)}, t.EFFECTIVE_FROM=DATE '{EFFECTIVE_FROM}',
    t.UPDATED_AT=SYSTIMESTAMP, t.UPDATED_BY='{CREATED_BY}', t.RECORD_VERSION=t.RECORD_VERSION+1
WHEN NOT MATCHED THEN INSERT (CALCULATION_RULE_ID,TIER_NO,TIER_NAME_FA,LOWER_BOUND,UPPER_BOUND,BOUND_UNIT_CODE,TIER_BASIS_CODE,TIER_STRATEGY_CODE,FIXED_AMOUNT,RATE_VALUE,EFFECTIVE_FROM,CREATED_BY)
VALUES (s.CALCULATION_RULE_ID,{tier_no},{sql_str(name_fa)},{sql_num(lower)},{sql_num(upper)},'IRR',{sql_str(tier_basis)},{sql_str(tier_strategy)},{sql_num(fixed)},{sql_num(rate)},DATE '{EFFECTIVE_FROM}','{CREATED_BY}');
"""


def write_sql(path: Path, items: list[LogicalTariff]) -> None:
    by_first = {x.source_rows[0]: x for x in items}
    out: list[str] = []
    out += [
        "-- ============================================================================\n",
        "-- CBI Fee Tariff 1404 - clean/idempotent import generated from cbi-fee-1404.xlsx\n",
        f"-- Source SHA256: {SOURCE_EXPECTED_SHA256}\n",
        "-- Physical source rows: 239; logical tariff definitions: 229.\n",
        "-- No source formula is invented: ambiguous/external rows use EXTERNAL_VALUE.\n",
        "-- ============================================================================\n",
        "SET DEFINE OFF;\nWHENEVER SQLERROR EXIT SQL.SQLCODE ROLLBACK;\nALTER SESSION SET CURRENT_SCHEMA = FEE;\n\n",
        "PROMPT [CBI1404] Validating prerequisite baseline/reference data ...\n",
        f"DECLARE v NUMBER; BEGIN SELECT COUNT(*) INTO v FROM FEE_REGULATORY_SOURCE WHERE SOURCE_CODE={sql_str(SOURCE_REG_CODE)}; IF v<>1 THEN RAISE_APPLICATION_ERROR(-20041,'CBI 1404 regulatory source missing/duplicated'); END IF; END;\n/\n",
        f"DECLARE v NUMBER; BEGIN SELECT COUNT(*) INTO v FROM FEE_POLICY_VERSION pv JOIN FEE_POLICY_SET ps ON ps.POLICY_SET_ID=pv.POLICY_SET_ID WHERE ps.POLICY_CODE={sql_str(POLICY_CODE)} AND pv.VERSION_NO={sql_str(POLICY_VERSION_NO)}; IF v<>1 THEN RAISE_APPLICATION_ERROR(-20042,'CBI 1404 policy version missing/duplicated'); END IF; END;\n/\n\n",
        "PROMPT [CBI1404] Upserting 17 source section features ...\n",
    ]
    for section, (short, feature_code) in SECTION_CODES.items():
        out.append(merge_feature(section, short, feature_code))

    out.append("\nPROMPT [CBI1404] Upserting 229 logical fee definitions ...\n")
    for item in items:
        out.append(merge_definition(item))

    out.append("\nPROMPT [CBI1404] Upserting 229 fee definition versions ...\n")
    for item in items:
        out.append(merge_version(item))

    out.append("\nPROMPT [CBI1404] Upserting source-derived calculation rules ...\n")
    existing_rows = set(EXISTING_FEE_CODE_BY_SOURCE_ROW)
    tier_first_rows = {group[0] for group in TIER_GROUPS}
    # New/non-enriched logical tariffs: one source-derived rule each (tier groups handled separately).
    for item in items:
        first = item.source_rows[0]
        if first in existing_rows or first in tier_first_rows:
            continue
        out.append(merge_calc_rule(item))

    # Existing rich source mappings: update/create their established rules using the clean source semantics.
    guarantee = by_first[3]
    out.append(merge_calc_rule(guarantee, "GUARANTEE_ANNUAL", strategy="ANNUALIZED_PERCENTAGE", basis="RateWithMinimumAmount",
                               rate=Decimal("0.005"), min_fee=Decimal("812500"), period="YEAR", name="محاسبه سالانه ضمانت نامه",
                               description=build_description(guarantee)))

    bill = by_first[42]
    out.append(merge_calc_rule(bill, "BILL_FIXED", strategy="FIXED", basis="Flat", fixed=Decimal("90000"),
                               name="کارمزد ثابت وصول بروات", description=build_description(bill)))
    out.append(merge_calc_rule(bill, "POSTAGE_EXTERNAL", strategy="EXTERNAL_VALUE", basis="Other", fixed=None, rate=None, min_fee=None, max_fee=None,
                               name="هزینه پست بیرونی", description="هزینه پست طبق متن منبع جداگانه و به صورت مقدار بیرونی اخذ می شود.", priority=20))

    commitment = by_first[118]
    out.append(merge_calc_rule(commitment, "COMMITMENT_ANNUAL", strategy="ANNUALIZED_PERCENTAGE", basis="Percentage",
                               rate=Decimal("0.01"), period="YEAR", name="کارمزد سالانه عدم استفاده از حد اعتباری",
                               description=build_description(commitment)))

    expertise = by_first[119]
    out.append(merge_calc_rule(expertise, "CREDIT_STAGE1", strategy="PERCENTAGE", basis="Percentage", rate=Decimal("0.0005"),
                               name="مرحله پذیرش درخواست", description="۰٫۵ در هزار مبلغ درخواستی در زمان پذیرش؛ غیرقابل برگشت.", priority=10))
    # Critical correction: the second stage is the remaining 1.0 per thousand so total = 1.5 per thousand, not an extra 1.5.
    out.append(merge_calc_rule(expertise, "CREDIT_STAGE2", strategy="PERCENTAGE", basis="Percentage", rate=Decimal("0.001"),
                               name="مرحله تصویب و انعقاد", description="مابقی کارمزد تا مجموع ۱٫۵ در هزار پس از کسر ۰٫۵ در هزار مرحله پذیرش.", priority=20))

    cheque = by_first[54]
    out.append(merge_calc_rule(cheque, "CHEQUEBOOK_FIXED", strategy="FIXED", basis="Flat", fixed=Decimal("75000"),
                               name="مبلغ پایه صدور دسته چک", description=build_description(cheque)))
    out.append(merge_calc_rule(cheque, "PRINTING_EXTERNAL", strategy="EXTERNAL_VALUE", basis="Other", name="هزینه چاپ بیرونی",
                               description="مبلغ پرداختی به چاپخانه دولتی/بهای تمام شده مطابق منبع و به صورت مقدار بیرونی.", priority=20))
    out.append(merge_calc_rule(cheque, "STAMP_EXTERNAL", strategy="EXTERNAL_VALUE", basis="Other", name="هزینه تمبر بیرونی",
                               description="هزینه تمبر مطابق منبع و به صورت مقدار بیرونی.", priority=30))

    out.append("\nPROMPT [CBI1404] Upserting five appraisal bracket models (3 tiers each) ...\n")
    for group, cfg in TIER_GROUPS.items():
        item = by_first[group[0]]
        rule_code = "APPRAISAL_COMPOSITE" if group[0] == 90 else "CBI1404_APPRAISAL_COMPOSITE"
        out.append(merge_calc_rule(item, rule_code, strategy="COMPOSITE", basis="RatePerBrackets", fixed=None, rate=None,
                                   name=f"محاسبه ترکیبی - {item.name_fa}", description=build_description(item)))
        out.append(merge_input(item.fee_code, rule_code, "APPRAISAL_AMOUNT", "مبلغ ارزیابی", 1))
        out.append(merge_input(item.fee_code, rule_code, "OFFICIAL_EXPERT_TARIFF", "تعرفه کارشناس رسمی", 2))
        out.append(merge_component(item.fee_code, rule_code, 1, "OPERATOR", "کمینه محاسبه بانکی و تعرفه کارشناس رسمی", operator="MIN_OF"))
        out.append(merge_component(item.fee_code, rule_code, 2, "OPERATOR", "محاسبه دو در هزار مازاد", parent_seq=1, operator="MULTIPLY"))
        out.append(merge_component(item.fee_code, rule_code, 3, "INPUT", "مبلغ ارزیابی", parent_seq=2, input_code="APPRAISAL_AMOUNT"))
        out.append(merge_component(item.fee_code, rule_code, 4, "CONSTANT", "نرخ دو در هزار", parent_seq=2, constant=Decimal("0.002")))
        out.append(merge_component(item.fee_code, rule_code, 5, "EXTERNAL_VALUE", "تعرفه کارشناس رسمی", parent_seq=1, reference="OFFICIAL_EXPERT_TARIFF"))
        first_upper = item.tier_first_upper
        assert first_upper is not None
        out.append(merge_tier(item.fee_code, rule_code, 1, f"تا {sql_num(first_upper)} ریال", Decimal("0"), first_upper,
                              "WHOLE_AMOUNT", "FIXED", fixed=Decimal("2587500")))
        out.append(merge_tier(item.fee_code, rule_code, 2, "مازاد تا ۴۰۰ میلیارد ریال", first_upper, Decimal("400000000000"),
                              "EXCESS_OVER_LOWER_BOUND", "COMPOSITE", rate=Decimal("0.002")))
        out.append(merge_tier(item.fee_code, rule_code, 3, "مازاد بر ۴۰۰ میلیارد ریال - ارجاع به کارشناس رسمی", Decimal("400000000000"), None,
                              "EXCESS_OVER_LOWER_BOUND", "EXTERNAL_VALUE"))

    out += [
        "\nPROMPT [CBI1404] DML staged. Do not COMMIT before verification.\n",
        "PROMPT [CBI1404] Run through 00-install-cbi-fee-1404.sql for verify-then-commit semantics.\n",
    ]
    path.write_text("".join(out), encoding="utf-8")


def write_verify_sql(path: Path, expected: dict[str, int]) -> None:
    logical = expected["logical_tariffs"]
    tiers = expected["official_tiers"]
    sql = f"""-- CBI Fee 1404 verification
SET DEFINE OFF;
WHENEVER SQLERROR EXIT SQL.SQLCODE ROLLBACK;
ALTER SESSION SET CURRENT_SCHEMA = FEE;
SET SERVEROUTPUT ON;

DECLARE
  v NUMBER;
  PROCEDURE assert_eq(p_label VARCHAR2, p_actual NUMBER, p_expected NUMBER) IS
  BEGIN
    DBMS_OUTPUT.PUT_LINE(RPAD(p_label,45)||' actual='||p_actual||' expected='||p_expected);
    IF p_actual <> p_expected THEN RAISE_APPLICATION_ERROR(-20140, p_label||' mismatch'); END IF;
  END;
BEGIN
  SELECT COUNT(*) INTO v FROM FEE_DEFINITION WHERE CLASSIFICATION_CODE='{CLASSIFICATION_CODE}' AND IS_ACTIVE='Y';
  assert_eq('CBI1404 active definitions', v, {logical});

  SELECT COUNT(*) INTO v
    FROM FEE_DEFINITION_VERSION dv
    JOIN FEE_DEFINITION d ON d.FEE_DEFINITION_ID=dv.FEE_DEFINITION_ID
    JOIN FEE_REGULATORY_SOURCE rs ON rs.REGULATORY_SOURCE_ID=dv.REGULATORY_SOURCE_ID
   WHERE d.CLASSIFICATION_CODE='{CLASSIFICATION_CODE}' AND dv.VERSION_NO='1.0' AND rs.SOURCE_CODE='{SOURCE_REG_CODE}';
  assert_eq('CBI1404 definition versions', v, {logical});

  SELECT COUNT(*) INTO v
    FROM FEE_DEFINITION d
   WHERE d.CLASSIFICATION_CODE='{CLASSIFICATION_CODE}'
     AND NOT EXISTS (SELECT 1 FROM FEE_DEFINITION_VERSION dv WHERE dv.FEE_DEFINITION_ID=d.FEE_DEFINITION_ID AND dv.VERSION_NO='1.0');
  assert_eq('definitions without version', v, 0);

  SELECT COUNT(*) INTO v
    FROM FEE_DEFINITION d JOIN FEE_DEFINITION_VERSION dv ON dv.FEE_DEFINITION_ID=d.FEE_DEFINITION_ID
   WHERE d.CLASSIFICATION_CODE='{CLASSIFICATION_CODE}'
     AND NOT EXISTS (SELECT 1 FROM FEE_CALCULATION_RULE r WHERE r.FEE_DEFINITION_VERSION_ID=dv.FEE_DEFINITION_VERSION_ID AND r.IS_ACTIVE='Y');
  assert_eq('definitions without active calc rule', v, 0);

  SELECT COUNT(*) INTO v
    FROM FEE_CALCULATION_RULE r
    JOIN FEE_DEFINITION_VERSION dv ON dv.FEE_DEFINITION_VERSION_ID=r.FEE_DEFINITION_VERSION_ID
    JOIN FEE_DEFINITION d ON d.FEE_DEFINITION_ID=dv.FEE_DEFINITION_ID
   WHERE d.CLASSIFICATION_CODE='{CLASSIFICATION_CODE}' AND r.RATE_VALUE IS NOT NULL AND (r.RATE_VALUE < 0 OR r.RATE_VALUE > 1);
  assert_eq('invalid normalized rate values', v, 0);

  SELECT COUNT(*) INTO v
    FROM FEE_CALCULATION_TIER t
    JOIN FEE_CALCULATION_RULE r ON r.CALCULATION_RULE_ID=t.CALCULATION_RULE_ID
    JOIN FEE_DEFINITION_VERSION dv ON dv.FEE_DEFINITION_VERSION_ID=r.FEE_DEFINITION_VERSION_ID
    JOIN FEE_DEFINITION d ON d.FEE_DEFINITION_ID=dv.FEE_DEFINITION_ID
   WHERE d.CLASSIFICATION_CODE='{CLASSIFICATION_CODE}';
  assert_eq('CBI1404 appraisal tiers', v, {tiers});

  SELECT COUNT(*) INTO v FROM (
    SELECT FEE_CODE FROM FEE_DEFINITION WHERE CLASSIFICATION_CODE='{CLASSIFICATION_CODE}' GROUP BY FEE_CODE HAVING COUNT(*)>1
  );
  assert_eq('duplicate internal fee codes', v, 0);

  SELECT COUNT(*) INTO v
    FROM FEE_CALCULATION_RULE r
    JOIN FEE_DEFINITION_VERSION dv ON dv.FEE_DEFINITION_VERSION_ID=r.FEE_DEFINITION_VERSION_ID
    JOIN FEE_DEFINITION d ON d.FEE_DEFINITION_ID=dv.FEE_DEFINITION_ID
   WHERE d.FEE_CODE='CBI_CREDIT_EXPERTISE_FEE' AND r.RULE_CODE='CREDIT_STAGE2' AND r.RATE_VALUE=0.001;
  assert_eq('corrected credit expertise stage2', v, 1);

  DBMS_OUTPUT.PUT_LINE('CBI Fee 1404 verification OK.');
END;
/

SELECT CALCULATION_STRATEGY_CODE, COUNT(*) AS RULE_COUNT
  FROM FEE_CALCULATION_RULE r
  JOIN FEE_DEFINITION_VERSION dv ON dv.FEE_DEFINITION_VERSION_ID=r.FEE_DEFINITION_VERSION_ID
  JOIN FEE_DEFINITION d ON d.FEE_DEFINITION_ID=dv.FEE_DEFINITION_ID
 WHERE d.CLASSIFICATION_CODE='{CLASSIFICATION_CODE}' AND r.IS_ACTIVE='Y'
 GROUP BY CALCULATION_STRATEGY_CODE
 ORDER BY CALCULATION_STRATEGY_CODE;
"""
    path.write_text(sql, encoding="utf-8")


def validate(items: list[LogicalTariff], source_rows: list[SourceRow]) -> dict[str, int]:
    assert len(source_rows) == 239, len(source_rows)
    assert len(items) == 229, len(items)
    assert len({x.fee_code for x in items}) == len(items)
    assert len({x.section for x in items}) == 17
    covered = sorted(r for x in items for r in x.source_rows)
    assert covered == list(range(2, 241)), (covered[:5], covered[-5:], len(covered))
    # 10 physical continuation rows were consolidated into 5 bracket tariffs.
    assert sum(len(x.source_rows) - 1 for x in items) == 10
    # Validate rate normalization and source completeness.
    for x in items:
        if x.rate_value is not None:
            assert Decimal("0") <= x.rate_value <= Decimal("1")
        if x.calc_strategy == "FIXED":
            assert x.fixed_amount is not None
        if x.calc_strategy == "COMPOSITE":
            assert x.tier_first_upper is not None
    assert any(x.import_status == "REVIEW_CONFLICT" for x in items)
    assert any(x.import_status == "EXTERNAL_RULE" for x in items)
    # Five official appraisal logical tariffs * 3 tiers = 15.
    counts = Counter(x.import_status for x in items)
    return {
        "physical_source_rows": len(source_rows),
        "logical_tariffs": len(items),
        "source_sections": len({x.section for x in items}),
        "consolidated_continuation_rows": 10,
        "official_tiers": 15,
        "safe": counts["SAFE"],
        "safe_tiered_external": counts["SAFE_TIERED_EXTERNAL"],
        "existing_enriched": counts["EXISTING_ENRICHED"],
        "external_rule": counts["EXTERNAL_RULE"],
        "review_conflict": counts["REVIEW_CONFLICT"],
        "review_source_gap": counts["REVIEW_SOURCE_GAP"],
    }


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("source", type=Path)
    ap.add_argument("outdir", type=Path)
    args = ap.parse_args()
    args.outdir.mkdir(parents=True, exist_ok=True)
    rows, sha = source_rows_from_workbook(args.source)
    if sha != SOURCE_EXPECTED_SHA256:
        raise SystemExit(f"Source SHA256 mismatch: got {sha}, expected {SOURCE_EXPECTED_SHA256}")
    items = build_logical(rows)
    counts = validate(items, rows)
    write_csv(args.outdir / "cbi_fee_1404_clean.csv", items)
    write_review_csv(args.outdir / "cbi_fee_1404_review.csv", items)
    write_sql(args.outdir / "01-import-cbi-fee-1404.sql", items)
    write_verify_sql(args.outdir / "02-verify-cbi-fee-1404.sql", counts)
    manifest = {
        "source_file": args.source.name,
        "source_sha256": sha,
        "source_sheet": "کل کارمزدها",
        **counts,
        "classification_code": CLASSIFICATION_CODE,
        "regulatory_source_code": SOURCE_REG_CODE,
        "policy_code": POLICY_CODE,
        "policy_version": POLICY_VERSION_NO,
        "rate_normalization": "Excel percent value / 100 => RATE_VALUE decimal fraction",
        "ambiguous_formula_policy": "EXTERNAL_VALUE; no invented executable formula",
    }
    (args.outdir / "cbi_fee_1404_manifest.json").write_text(json.dumps(manifest, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps(manifest, ensure_ascii=False, indent=2))

if __name__ == "__main__":
    main()
